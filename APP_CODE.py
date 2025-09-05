import streamlit as st
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xlcalculator import ModelCompiler, Evaluator

st.set_page_config(page_title="BI UL — Benefit Illustration (no VBA)", layout="wide")
st.title("📋 Benefit Illustration — Streamlit (Excel formulas only, VBA ignored)")

# Helper: load workbook (from upload or local fallback)
def load_workbook_from_upload(uploaded_file):
    if uploaded_file is None:
        # Try fallback local path used in this session if present
        fallback = "/mnt/data/BI UL.xlsm"
        if os.path.exists(fallback):
            return load_workbook(fallback, data_only=False)
        return None
    else:
        # Write upload to temp file and load
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[-1])
        tmp.write(uploaded_file.getbuffer())
        tmp.flush()
        tmp.close()
        wb = load_workbook(tmp.name, data_only=False)
        # keep temp file path for later saving
        return wb, tmp.name

# Extract defined names (named ranges) that point to single cells and are on the Input sheet
def get_named_inputs_from_wb(wb, sheet_name_candidates=("Input", "INPUT", "input")):
    inputs = {}
    dn = wb.defined_names
    # iterate defined names
    for defn in dn.definedName:
        name = defn.name
        # skip built-ins
        if name is None or name.startswith('_xlnm'):
            continue
        # get destinations (sheet, coord) pairs; may contain external links
        try:
            destinations = list(defn.destinations)
        except Exception:
            continue
        # pick destinations that belong to Input-like sheets
        for (sname, coord) in destinations:
            if sname in wb.sheetnames and any(sname == c for c in sheet_name_candidates):
                # We only handle named ranges that map to a single cell
                if ":" not in coord:
                    inputs[name] = (sname, coord)
                else:
                    # if range but single-cell range like 'A2:A2', treat it
                    # else skip multi-cell ranges
                    a, b = coord.split(":")
                    if a == b:
                        inputs[name] = (sname, a)
                break
    return inputs

# Fallback heuristic: assume Input sheet is key-value table with first column labels and second column default values
def get_key_value_inputs_from_sheet(wb, sheet_name_candidates=("Input", "INPUT", "input")):
    for candidate in sheet_name_candidates:
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            # read up to first 200 rows
            kv = {}
            for r in range(1, min(200, ws.max_row) + 1):
                key_cell = ws.cell(row=r, column=1).value
                val_cell = ws.cell(row=r, column=2).value
                if key_cell is None:
                    continue
                # only accept reasonable keys
                key = str(key_cell).strip()
                if key == "":
                    continue
                kv[key] = {"sheet": candidate, "cell": f"{get_column_letter(2)}{r}", "default": val_cell}
            return kv
    return {}

# Write inputs (dict of (sheet, cell)->value) into a workbook and save to a temp file, return temp path
def write_inputs_to_tempfile_and_save(wb, inputs_map, original_tmp_path=None):
    # wb is an openpyxl workbook object
    # inputs_map: dict with key either named (name->(sheet, coord)) or direct mapping keys->value as (sheet, cell)
    # If original_tmp_path provided, overwrite it; else create temp file
    tmpfile = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmpfile.close()
    # apply inputs_map
    for k, v in inputs_map.items():
        # v is (sheet, cell, value) OR (sheet, cell) with value in 'value' in tuple
        if isinstance(v, tuple) and len(v) == 3:
            sname, coord, val = v
            if sname in wb.sheetnames:
                ws = wb[sname]
                ws[coord] = val
        elif isinstance(v, dict) and {"sheet", "cell", "value"}.issubset(v.keys()):
            sname = v["sheet"]; coord = v["cell"]; val = v["value"]
            if sname in wb.sheetnames:
                ws = wb[sname]
                ws[coord] = val
    # Save workbook as a temporary file (xlcalculator expects a real file path)
    wb.save(tmpfile.name)
    return tmpfile.name

# Evaluate all cells on Output sheet and return as dataframe + dict of key outputs (non-empty cells)
def evaluate_output_sheet(tempfile_path, output_sheet_name_guess=("Output", "OUTPUT", "output")):
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(tempfile_path)
    evaluator = Evaluator(model)

    # find output sheet
    sheetname = None
    for s in output_sheet_name_guess:
        if s in model.workbook.sheets:
            sheetname = s
            break
    # fallback: take first sheet that equals 'Output' ignoring case
    if sheetname is None:
        for s in model.workbook.sheets:
            if s.lower() == "output":
                sheetname = s
                break
    if sheetname is None:
        # if still none, pick any sheet named 'Output' with case-insensitive match
        for s in model.workbook.sheets:
            if "output" in s.lower():
                sheetname = s
                break
    if sheetname is None:
        # if still none, pick last sheet as fallback
        sheetname = list(model.workbook.sheets.keys())[-1]

    # Get sheet size via openpyxl to iterate
    wb2 = load_workbook(tempfile_path, data_only=False)
    ws2 = wb2[sheetname]

    rows = []
    outputs = {}
    for r in range(1, ws2.max_row + 1):
        row_vals = []
        any_nonempty = False
        for c in range(1, ws2.max_column + 1):
            cell_coord = f"{get_column_letter(c)}{r}"
            ref = f"'{sheetname}'!{cell_coord}"
            try:
                val = evaluator.evaluate(ref)
            except Exception:
                # fallback: read the value from the saved workbook (may be plain values)
                val = ws2.cell(row=r, column=c).value
            row_vals.append(val)
            if val is not None and (str(val).strip() != ""):
                any_nonempty = True
        if any_nonempty:
            rows.append(row_vals)
    # build dataframe with column headers if first row appears to be header-like
    df = pd.DataFrame(rows)
    return df, sheetname

# UI: Upload / load workbook
uploaded_file = st.file_uploader("Upload BI UL Excel file (.xlsm or .xlsx). If none uploaded the app will try `/mnt/data/BI UL.xlsm`.", type=["xlsm", "xlsx", "xls"])

wb_obj = None
orig_tmp_path = None
if uploaded_file is not None:
    try:
        wb_obj, orig_tmp_path = load_workbook_from_upload(uploaded_file)
    except Exception as e:
        st.error(f"Failed to read uploaded workbook: {e}")
        st.stop()
else:
    # try fallback local
    fallback = "/mnt/data/BI UL.xlsm"
    if os.path.exists(fallback):
        try:
            wb_obj = load_workbook(fallback, data_only=False)
            orig_tmp_path = fallback
            st.info(f"Loaded fallback workbook from {fallback}")
        except Exception as e:
            st.error(f"Failed to load fallback workbook {fallback}: {e}")
            st.stop()
    else:
        st.info("Please upload your BI UL.xlsm file to continue.")
        st.stop()

# At this point wb_obj is an openpyxl workbook
# Try to detect named inputs
named_inputs = get_named_inputs_from_wb(wb_obj)
use_named = len(named_inputs) > 0

st.sidebar.header("Input mode")
if use_named:
    st.sidebar.success(f"Detected {len(named_inputs)} named input(s) on Input sheet. Using named inputs.")
else:
    st.sidebar.warning("No named inputs detected. Falling back to key-value heuristic from the Input sheet.")

inputs_map_for_writing = {}  # will contain mapping name->(sheet, cell, value) to be written

if use_named:
    st.sidebar.subheader("Named inputs (edit values)")
    for name, (sheet, coord) in named_inputs.items():
        # get current default value from workbook
        try:
            # read cell value (current)
            val = wb_obj[sheet][coord].value if coord in wb_obj[sheet] else wb_obj[sheet][coord].value
        except Exception:
            # safer: parse coordinate to numeric
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            try:
                (col, row) = coordinate_from_string(coord)
                val = wb_obj[sheet].cell(row=row, column=column_index_from_string(col)).value
            except Exception:
                val = None
        # create widget based on type
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).replace('.', '', 1).isdigit()):
            new_val = st.sidebar.number_input(f"{name} ({sheet}!{coord})", value=float(val) if val not in (None, "") else 0.0)
        else:
            new_val = st.sidebar.text_input(f"{name} ({sheet}!{coord})", value=str(val) if val is not None else "")
        inputs_map_for_writing[name] = (sheet, coord, new_val)
else:
    kv = get_key_value_inputs_from_sheet(wb_obj)
    if not kv:
        st.warning("Could not detect Input sheet or key-value pairs. You can still manually provide inputs below.")
        # allow user to enter arbitrary inputs and map them to known cells if desired
        st.sidebar.subheader("Manual inputs")
        manual_inputs = {}
        n = st.sidebar.number_input("How many manual input fields to add?", min_value=0, max_value=50, value=0)
        for i in range(n):
            label = st.sidebar.text_input(f"Label #{i+1}", value=f"Input_{i+1}")
            sheet_name = st.sidebar.text_input(f"Sheet for {label}", value="Input")
            cell = st.sidebar.text_input(f"Cell for {label}", value=f"A{i+2}")
            value = st.sidebar.text_input(f"Value for {label}", value="")
            if label:
                inputs_map_for_writing[label] = {"sheet": sheet_name, "cell": cell, "value": value}
    else:
        st.sidebar.subheader("Detected Input key→value pairs (edit defaults)")
        for key, meta in kv.items():
            default = meta.get("default", "")
            if isinstance(default, (int, float)):
                new_val = st.sidebar.number_input(f"{key} ({meta['sheet']}!{meta['cell']})", value=float(default))
            else:
                # if it's empty numeric? allow numeric too
                try:
                    new_val = float(default) if default not in (None, "") and str(default).replace('.', '', 1).isdigit() else st.sidebar.text_input(f"{key} ({meta['sheet']}!{meta['cell']})", value=str(default) if default is not None else "")
                except Exception:
                    new_val = st.sidebar.text_input(f"{key} ({meta['sheet']}!{meta['cell']})", value=str(default) if default is not None else "")
            inputs_map_for_writing[key] = {"sheet": meta["sheet"], "cell": meta["cell"], "value": new_val}

# Build the mapping in the format (sheet, cell, value)
final_inputs = {}
for k, v in inputs_map_for_writing.items():
    if isinstance(v, tuple) and len(v) == 3:
        sheet, coord, val = v
        final_inputs[k] = (sheet, coord, val)
    elif isinstance(v, dict) and {"sheet", "cell", "value"}.issubset(v.keys()):
        final_inputs[k] = (v["sheet"], v["cell"], v["value"])
    else:
        # fallback - skip
        continue

# Confirm & compute button
if st.sidebar.button("Compute Benefit Illustration"):
    with st.spinner("Writing inputs and evaluating formulas..."):
        try:
            # Use a fresh copy of the workbook to write inputs into and save as temp file
            # reload original to avoid overwriting
            if orig_tmp_path and os.path.exists(orig_tmp_path) and uploaded_file is not None:
                # if uploaded_file was provided, orig_tmp_path was a temp file; we'll reload from it
                wb_to_write = load_workbook(orig_tmp_path, data_only=False)
            else:
                # reload from in-memory workbook object (wb_obj)
                wb_to_write = wb_obj

            temp_path = write_inputs_to_tempfile_and_save(wb_to_write, final_inputs, original_tmp_path=orig_tmp_path)
            df_out, out_sheet = evaluate_output_sheet(temp_path)
        except Exception as e:
            st.error(f"Computation failed: {e}")
            st.stop()

    st.success("Computation finished — showing Benefit Illustration outputs.")

    # Display summary metrics: Show top-left region or named outputs if any
    st.header("Final Benefit Illustration — Output sheet")
    st.subheader(f"Output sheet used: {out_sheet}")

    # Present the dataframe read from output sheet
    st.dataframe(df_out, use_container_width=True)

    # Additionally: try to detect key output cells (first column as label, second as numeric)
    if not df_out.empty:
        try:
            # try to interpret first two columns as label/value
            candidate = df_out.iloc[:, :2].dropna(how="all")
            candidate.columns = ["Label", "Value"] if candidate.shape[1] >= 2 else ["Label"]
            if "Value" in candidate.columns:
                st.subheader("Key outputs (interpreted from leftmost two columns)")
                st.table(candidate.head(50).set_index("Label"))
        except Exception:
            pass

    # allow download of the temp workbook with inputs written
    with open(temp_path, "rb") as f:
        st.download_button("Download workbook with inputs (temp file)", data=f, file_name="BI_UL_with_inputs.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.info("Note: Complex Excel functions or external links may not be fully supported by xlcalculator. This tool replicates workbook formula logic as best as xlcalculator supports and uses the workbook with input values written to produce evaluated outputs.")
else:
    st.sidebar.write("Press **Compute Benefit Illustration** to run formulas and show the Output sheet.")

# Final note
st.markdown(
    """
    **Limitations & notes**
    - This app ignores VBA/macros (as requested).
    - It relies on `xlcalculator` to evaluate Excel formulas. Some complex or Excel-specific functions may not be supported.
    - Best results occur when the workbook uses **named ranges** for inputs (the app tries to detect these). The fallback key→value heuristic expects the `Input` sheet to be a two-column list (labels in col A, defaults in col B).
    - If you want me to refine the mapping (e.g., point a named input to a different cell, or hard-code a few critical cell addresses), tell me which named ranges or Input sheet labels you want used and I can embed them directly in the app.
    """
)
